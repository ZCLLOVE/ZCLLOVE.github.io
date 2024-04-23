import openpyxl
import re

# UUID的正则表达式（32位UUID带连字符）
uuid_pattern_hyphen = re.compile(r'^[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12}$')
# UUID的正则表达式（32位UUID不带连字符）
uuid_pattern_no_hyphen = re.compile(r'^[a-fA-F0-9]{32}$')
# 分类的正则表达式（两个英文字母）
category_pattern = re.compile(r'^[A-Za-z]{2}$')

def collect_uuids_and_categories_from_excel(file_path):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)
    uuid_category_list = []
    uuid_set = set()  # 用于存储已经找到的UUID，避免重复

    # 遍历所有的工作表，从后往前
    for sheet in reversed(workbook.sheetnames):
        worksheet = workbook[sheet]
        # 从最后一行到第一行遍历
        for row in reversed(list(worksheet.iter_rows(values_only=True))):
            uuid = None
            category = None
            # 遍历行中的每一个单元格，寻找UUID和分类
            for cell_value in row:
                if cell_value and isinstance(cell_value, str):
                    # 如果单元格符合UUID格式
                    if uuid_pattern_hyphen.match(cell_value) or uuid_pattern_no_hyphen.match(cell_value):
                        uuid = cell_value
                    # 如果单元格符合分类格式（两个英文字母）
                    elif category_pattern.match(cell_value):
                        category = cell_value
            # 如果同时找到了UUID和分类，并且UUID未重复，收集起来
            if uuid and category and uuid not in uuid_set:
                uuid_category_list.append((category, uuid, sheet))
                uuid_set.add(uuid)  # 将UUID加入集合中，防止重复

    # 打印收集到的UUID和分类的SQL语句
    if uuid_category_list:
        print("Found UUIDs and Categories:")
        for category, uuid, sheet in uuid_category_list:
            # 打印插入的SQL语句，增加sheet信息
            print(f'INSERT INTO "1022review_template" (category, uuid, sheet_name) VALUES (\'{category}\', \'{uuid}\', \'{sheet}\');')
    else:
        print("No UUIDs or categories found.")

# 使用文件路径运行
collect_uuids_and_categories_from_excel(r'C:\Users\29510\Desktop\共享流程收集对比校验_1022_v3.xlsx')
